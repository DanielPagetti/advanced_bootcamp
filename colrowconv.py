import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string


print (get_column_letter(1))

get_column_letter(2)

get_column_letter(27)

get_column_letter(900)

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
get_column_letter(sheet.max_column)

column_index_from_string('A')

column_index_from_string('AA')
