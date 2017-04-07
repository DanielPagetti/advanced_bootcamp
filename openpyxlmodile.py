import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
type(wb)
sheet = wb.get_sheet_by_name('Sheet1')

sheet['A1']

sheet['A1'].value

c = sheet['B1']
c.value
'Row ' + str(c.row) + ', Column ' + c.column + ' is ' + c.value

'Cell ' + c.coordinate + ' is ' + c.value

sheet['C1'].value
sheet.cell(row=1, column=2)

sheet.cell(row=1, column=2).value

for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

type(sheet)
sheet.title
anotherSheet = wb.active
anotherSheet
